
from deep_translator import GoogleTranslator
from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import json


def get_ingredients_from_JSON_file(ingredients_file):
    with open(ingredients_file) as jsonfile:
        json_data = json.load(jsonfile)

    recipe = list()
    for key in json_data.keys():
        ingredients = pd.json_normalize(json_data[key])
        ingredients.loc[-1] = [key, "", ""]
        ingredients.index = ingredients.index + 1
        ingredients.sort_index(inplace=True)
        # print(ingredients['item'])
        recipe.append(ingredients)
    result = pd.concat(recipe)
    # print(result)
    result.reset_index()
    return result


def save_recipe_ingredients_in_worksheet(ingredients_df, recipe_ws, workbook):
    for r in dataframe_to_rows(ingredients_df, index=False, header=True):
        recipe_ws.append(r)
    workbook.save('Precios Pasteles.xlsx')


def translate_item(item):
    translated_text = GoogleTranslator(
        source='en', target='es').translate(item)
    return translated_text


def compare_item_with_raw_materials_list(item, raw_materials_list):
    model = SentenceTransformer('all-MiniLM-L6-v2')
    # raw_materials = pd.read_excel('Precios Pasteles.xlsx', sheet_name='Materia Prima')
    embed1 = model.encode(item)
    # raw_materials_list = raw_materials['Producto'].values.tolist()
    embed_raw_materials_list = model.encode(raw_materials_list)
    # print(raw_materials_list)
    similarities_list = cosine_similarity([embed1], embed_raw_materials_list)
    # most_similar_items = similarities_list[similarities_list >= 0.5]
    # print(similarities_list)
    indexes = np.where(similarities_list >= 0.55)[1].tolist()
    # print(most_similar_items, type(np.where(similarities_list >= 0.5)[1]))
    similar_items_list = [raw_materials_list[i] for i in indexes]
    print(type(indexes))
    return indexes, similar_items_list


def create_recipe_worksheet(recipe_name):
    workbook = load_workbook(filename="Precios Pasteles.xlsx")
    ws = workbook.create_sheet(recipe_name)
    return ws, workbook


def calculate_formula(similar_item, sim_items_dict, index, recipe_ws, workbook):
    sim_item_index = sim_items_dict[similar_item]+2
    ingredient_row = f"D{index+2}"
    quantity = f"B{index+2}"
    r = recipe_ws[quantity]
    recipe_ws[quantity] = int(r.value)
    recipe_ws[f"E{index+2}"] = f"""=IF('Materia Prima'!E{sim_item_index}="Kg",1/1000,1)"""
    formula = f"('Materia Prima'!F{sim_item_index}/'Materia Prima'!D{sim_item_index})*{quantity}"
    # print(r.value, type(r.value))
    recipe_ws[ingredient_row] = f"""=IF('Materia Prima'!E{sim_item_index}="Kg",{formula}/1000,{formula})"""
    # print(ingredient_row, recipe_ws[ingredient_row])
    workbook.save('Precios Pasteles.xlsx')


def main():
    item = 'whole wheat flour'
    translated_item = translate_item(item)
    # print(translated_item)
    # compare_item_with_raw_material(translated_item)


if __name__ == "__main__":
    main()
