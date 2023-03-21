
from deep_translator import GoogleTranslator
from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity
# import pandas as pd
import numpy as np
from openpyxl import load_workbook


def translate_item(item):

    translated_text = GoogleTranslator(
        source='en', target='es').translate(item)
    return translated_text


def compare_item_with_raw_materials_list(item, raw_materials_list):
    model = SentenceTransformer('all-MiniLM-L6-v2')
    # raw_materials = pd.read_excel('Precios Pasteles.xlsx', sheet_name='Materia Prima')
    embed1 = model.encode(item)
    # raw_materials_list = raw_materials['Producto'].values.tolist()
    embed_raw_materials_list = model.encode(raw_materials_list)
    # print(raw_materials_list)
    similarities_list = cosine_similarity([embed1], embed_raw_materials_list)
    # most_similar_items = similarities_list[similarities_list >= 0.5]
    # print(similarities_list)
    indexes = np.where(similarities_list >= 0.55)[1].tolist()
    # print(most_similar_items, type(np.where(similarities_list >= 0.5)[1]))
    similar_items_list = [raw_materials_list[i] for i in indexes]
    print(type(indexes))
    return indexes, similar_items_list


def create_recipe_worksheet(recipe_name):
    workbook = load_workbook(filename="Precios Pasteles.xlsx")
    ws = workbook.create_sheet(recipe_name)
    return ws, workbook


def append_row_to_recipe_worksheet(row, recipe_ws):
    print(type(row))
    # recipe_ws.append(row)


def handle_formulas(similar_item, sim_items_dict, index, recipe_ws, workbook):

    sim_item_index = sim_items_dict[similar_item]+2

    ingredient_row = f"D{index+2}"
    recipe_ws[ingredient_row] = f"=('Materia Prima'!F{sim_item_index}/'Materia Prima'!D{sim_item_index})"
    print(ingredient_row, recipe_ws[ingredient_row])
    workbook.save('Precios Pasteles.xlsx')


def main():
    item = 'whole wheat flour'
    translated_item = translate_item(item)
    # print(translated_item)
    # compare_item_with_raw_material(translated_item)


if __name__ == "__main__":
    main()
